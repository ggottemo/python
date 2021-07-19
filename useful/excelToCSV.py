#! python3
#excelToCSV.py - Convert excel files in current dir to csv files with the format 
#<excel filename>_<sheet title>

import os, csv, logging
from openpyxl import Workbook, load_workbook

logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s -  %(levelname)\
s -  %(message)s')
logging.disable()
#Create directory for new files
os.makedirs('Excel_CSV_Conversions', exist_ok=True)
for excelFile in os.listdir('.'):
    #skip non-xlsx files, load workbook object
    if not excelFile.endswith('.xlsx'):
        continue
    wb = load_workbook(excelFile)
    for sheetName in wb: #Loop through each sheet in the workbook
        sheet = wb[sheetName.title]
        
        #Create CSV filename from the Excel filename and sheet title
        fileName = str(excelFile)[:-5] + '_' + sheetName.title + '.csv'
        
        #Create csv.writer obj for this file
        csvFileObj = open(os.path.join('Excel_CSV_Conversions',fileName), 'w',\
                          newline='')
        writer = csv.writer(csvFileObj)
        print('Converting file to ' + fileName + '...')
        #loop through every row in the sheet
        
        logging.debug('Max Rows = %s'%(sheet.max_row) )
        for rowNum in range(1, sheet.max_row + 1):
            rowData =[] #List of cells
            #Loop through each cell in the row
            logging.debug('Row Loop!')
            if rowNum == 100 :
                break
            for colNum in range(1, sheet.max_column + 1 ):
                rowData.append(sheet.cell(rowNum, colNum).value)
                logging.debug('Column Loop!')
                
                
            #Write completed row data to a CSV row
            writer.writerow(rowData)
            
        #Close file
        csvFileObj.close()
        
print('Done.')